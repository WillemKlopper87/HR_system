from __future__ import annotations

import csv
import io
import xml.etree.ElementTree as ET

from openpyxl import Workbook
from openpyxl.styles import Font
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet

from .constants import AGGREGATE_ROW_KEYS, DEMOGRAPHIC_COLUMNS, OCCUPATIONAL_LEVEL_CODES, SKILLS_DEMOGRAPHIC_COLUMNS

# Reasonable, honest scope note: these exports reproduce every officially-
# required data point in a clearly labelled, spec-aligned structure (same
# row/column semantics as the EEA2/EEA4 forms — EEA-Form-Spec-Notes.md),
# not a pixel-perfect recreation of the department's exact merged-cell
# layout. A designated employer would still transcribe/upload this onto
# the actual online submission form or the .docx template.

_MATRIX_SECTIONS_EEA2 = [
    ("workforce_profile", "Workforce Profile (Section B)"),
    ("disability_workforce", "Employees with Disabilities (Section B)"),
    ("recruitment", "Recruitment (Section C)"),
    ("promotion", "Promotion (Section C)"),
    ("termination", "Termination (Section C)"),
]
_SKILLS_SECTION_EEA2 = ("skills_development", "Skills Development (Section D)")


def _row_labels():
    from core_hr.models import OccupationalLevel

    labels = {ol.code: ol.name for ol in OccupationalLevel.objects.filter(code__in=OCCUPATIONAL_LEVEL_CODES)}
    labels.update({
        "total_permanent": "Total permanent", "temporary_employees": "Temporary employees", "grand_total": "Grand total",
    })
    return labels


def _column_labels(columns):
    pretty = {
        "african_male": "African M", "coloured_male": "Coloured M", "indian_male": "Indian M", "white_male": "White M",
        "african_female": "African F", "coloured_female": "Coloured F", "indian_female": "Indian F", "white_female": "White F",
        "foreign_national_male": "Foreign National M", "foreign_national_female": "Foreign National F",
    }
    return [pretty[c] for c in columns]


def _matrix_rows(matrix: dict, columns) -> list[list]:
    row_labels = _row_labels()
    rows = [["Occupational level", *_column_labels(columns), "Total"]]
    for row_key in OCCUPATIONAL_LEVEL_CODES + AGGREGATE_ROW_KEYS:
        values = [matrix.get(row_key, {}).get(c, 0) for c in columns]
        rows.append([row_labels.get(row_key, row_key), *values, sum(values)])
    return rows


def to_csv(report) -> str:
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow([report.get_form_type_display(), report.report_year, f"v{report.version}", report.get_status_display()])
    writer.writerow([])

    if report.form_type == "eea2":
        for key, title in _MATRIX_SECTIONS_EEA2:
            writer.writerow([title])
            writer.writerows(_matrix_rows(report.data.get(key, {}), DEMOGRAPHIC_COLUMNS))
            writer.writerow([])
        key, title = _SKILLS_SECTION_EEA2
        writer.writerow([title])
        writer.writerows(_matrix_rows(report.data.get(key, {}), SKILLS_DEMOGRAPHIC_COLUMNS))
    else:
        writer.writerow(["Number of employees (Section C)"])
        writer.writerows(_matrix_rows(report.data.get("number_of_employees", {}), DEMOGRAPHIC_COLUMNS))
        writer.writerow([])
        writer.writerow(["Total remuneration (Section C)"])
        writer.writerows(_matrix_rows(report.data.get("total_remuneration", {}), DEMOGRAPHIC_COLUMNS))
        writer.writerow([])
        gap = report.data.get("median_and_gap", {})
        writer.writerow(["Median and gap statistics (Section E)"])
        for k, v in gap.items():
            writer.writerow([k, v])

    return buffer.getvalue()


def to_excel(report) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    def _write_sheet(title, rows):
        ws = wb.create_sheet(title[:31])
        for row in rows:
            ws.append(row)
        for cell in ws[1]:
            cell.font = Font(bold=True)

    cover = wb.create_sheet("Cover")
    cover.append([report.get_form_type_display(), report.report_year, f"Version {report.version}", report.get_status_display()])
    cover.append(["Period", str(report.period_start), str(report.period_end)])
    for cell in cover[1]:
        cell.font = Font(bold=True)

    if report.form_type == "eea2":
        for key, title in _MATRIX_SECTIONS_EEA2:
            _write_sheet(title, _matrix_rows(report.data.get(key, {}), DEMOGRAPHIC_COLUMNS))
        key, title = _SKILLS_SECTION_EEA2
        _write_sheet(title, _matrix_rows(report.data.get(key, {}), SKILLS_DEMOGRAPHIC_COLUMNS))
    else:
        _write_sheet("Number of employees", _matrix_rows(report.data.get("number_of_employees", {}), DEMOGRAPHIC_COLUMNS))
        _write_sheet("Total remuneration", _matrix_rows(report.data.get("total_remuneration", {}), DEMOGRAPHIC_COLUMNS))
        gap_rows = [["Metric", "Value"]] + [[k, str(v)] for k, v in report.data.get("median_and_gap", {}).items()]
        _write_sheet("Median and gap", gap_rows)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def to_pdf(report) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=15 * mm, bottomMargin=15 * mm)
    styles = getSampleStyleSheet()
    story = [
        Paragraph(f"{report.get_form_type_display()} — {report.report_year} (v{report.version})", styles["Title"]),
        Paragraph(f"Status: {report.get_status_display()} | Period: {report.period_start} to {report.period_end}", styles["Normal"]),
        Spacer(1, 10 * mm),
    ]

    def _add_table(title, rows):
        story.append(Paragraph(title, styles["Heading2"]))
        table = Table(rows, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f2937")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        story.append(table)
        story.append(Spacer(1, 6 * mm))

    if report.form_type == "eea2":
        for key, title in _MATRIX_SECTIONS_EEA2:
            _add_table(title, _matrix_rows(report.data.get(key, {}), DEMOGRAPHIC_COLUMNS))
        key, title = _SKILLS_SECTION_EEA2
        _add_table(title, _matrix_rows(report.data.get(key, {}), SKILLS_DEMOGRAPHIC_COLUMNS))
    else:
        _add_table("Number of employees", _matrix_rows(report.data.get("number_of_employees", {}), DEMOGRAPHIC_COLUMNS))
        _add_table("Total remuneration", _matrix_rows(report.data.get("total_remuneration", {}), DEMOGRAPHIC_COLUMNS))
        gap = report.data.get("median_and_gap", {})
        _add_table("Median and gap statistics", [["Metric", "Value"], *[[k, str(v)] for k, v in gap.items()]])

    doc.build(story)
    return buffer.getvalue()


def _dict_to_xml(parent: ET.Element, data) -> None:
    if isinstance(data, dict):
        for key, value in data.items():
            child = ET.SubElement(parent, _xml_safe_tag(key))
            _dict_to_xml(child, value)
    elif isinstance(data, list):
        for item in data:
            child = ET.SubElement(parent, "item")
            _dict_to_xml(child, item)
    else:
        parent.text = "" if data is None else str(data)


def _xml_safe_tag(key: str) -> str:
    return key if key and key[0].isalpha() else f"_{key}"


def to_xml(report) -> str:
    root = ET.Element(report.form_type)
    root.set("year", str(report.report_year))
    root.set("version", str(report.version))
    root.set("status", report.status)
    _dict_to_xml(root, report.data)
    return ET.tostring(root, encoding="unicode")
